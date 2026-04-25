"""
Fibonacci Retracement Analysis using Zerodha Kite API
======================================================
Input : Excel file with columns: date, symbol
Output: Color-coded Excel file with full Fibonacci + entry/SL analysis

Logic:
- 15-min range candle : first candle at 09:15
- Fibonacci levels    : 38.2%, 50%, 61.8% retracement from 15-min high downward
- Long entry trigger  : 15-min high + 1
- Stop loss           : low of the 5-min entry candle - 1
- Max gain / Max gain%: highest price reached after entry BEFORE SL is hit
- Checks if price retraced to each Fib level in subsequent 5-min candles
- Finds entry candle, SL hit candle, max gain, max gain %

Setup:
    pip install kiteconnect pandas openpyxl

Usage:
    python fib_analysis.py
    python fib_analysis.py --input my_file.xlsx --output results.xlsx
"""

import argparse
import sys
import time

import pandas as pd
from kiteconnect import KiteConnect
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG  ← update these before running
# ─────────────────────────────────────────────────────────────────────────────
API_KEY      = "kb7ux8tgqc6kek6c"
API_SECRET   = "nncu2lwokkz206tjz863ci6uka4voucs"
ACCESS_TOKEN = "kb7ux8tgqc6kek6c:PUuB66rIQX69NuTv18TUuIaDmkhbrsWX"   # leave as-is to trigger interactive login

EXCHANGE     = "NSE"
MARKET_OPEN  = "09:15"
FIB_LEVELS   = [0.382, 0.500, 0.618]
API_DELAY    = 0.35   # seconds between API calls to respect rate limits
# ─────────────────────────────────────────────────────────────────────────────


# ── Kite connection ───────────────────────────────────────────────────────────

def get_kite() -> KiteConnect:
    kite = KiteConnect(api_key=API_KEY)
    if ACCESS_TOKEN and ACCESS_TOKEN != "YOUR_ACCESS_TOKEN":
        kite.set_access_token(ACCESS_TOKEN)
    else:
        print("Login URL:\n", kite.login_url(), "\n")
        req_token = input("Paste request_token from redirect URL: ").strip()
        session   = kite.generate_session(req_token, api_secret=API_SECRET)
        kite.set_access_token(session["access_token"])
        print(f"Access token (save for next run): {session['access_token']}\n")
    return kite

# ── Instrument token lookup ───────────────────────────────────────────────────

_token_cache: dict = {}

def get_token(kite: KiteConnect, symbol: str, exchange: str = EXCHANGE) -> int:
    key = f"{exchange}:{symbol}"
    if key in _token_cache:
        return _token_cache[key]
    instruments = kite.instruments(exchange)
    for inst in instruments:
        if inst["tradingsymbol"] == symbol and inst["instrument_type"] == "EQ":
            _token_cache[key] = inst["instrument_token"]
            return inst["instrument_token"]
    raise ValueError(f"Instrument not found: {key}")


# ── Historical data ───────────────────────────────────────────────────────────

def fetch_candles(kite: KiteConnect, token: int, date_str: str, interval: str) -> pd.DataFrame:
    from_dt = f"{date_str} 09:15:00"
    to_dt   = f"{date_str} 15:30:00"
    try:
        data = kite.historical_data(token, from_dt, to_dt, interval)
    except Exception as e:
        print(f"    [WARN] {e}")
        return pd.DataFrame()
    if not data:
        return pd.DataFrame()
    df = pd.DataFrame(data)
    df["date"] = pd.to_datetime(df["date"])
    return df.sort_values("date").reset_index(drop=True)


# ── Fibonacci helpers ─────────────────────────────────────────────────────────

def fib_price(high: float, low: float, level: float) -> float:
    """Retracement level from high downward (long setup)."""
    return round(high - level * (high - low), 2)


def first_touch(df: pd.DataFrame, price: float) -> str:
    """HH:MM of first candle whose range contains price, else 'Not touched'."""
    for _, row in df.iterrows():
        if row["low"] <= price <= row["high"]:
            return row["date"].strftime("%H:%M")
    return "Not touched"


# ── Core analysis ─────────────────────────────────────────────────────────────

def analyse(kite: KiteConnect, symbol: str, date_str: str) -> dict:
    base = {"date": date_str, "symbol": symbol}

    # 1. Instrument token
    try:
        token = get_token(kite, symbol)
    except ValueError as e:
        return {**base, "error": str(e)}

    # 2. First 15-min candle at 09:15
    df15 = fetch_candles(kite, token, date_str, "15minute")
    time.sleep(API_DELAY)
    if df15.empty:
        return {**base, "error": "No 15-min data"}

    c15 = df15[df15["date"].dt.strftime("%H:%M") == MARKET_OPEN]
    if c15.empty:
        return {**base, "error": "09:15 candle not found"}
    c15 = c15.iloc[0]

    high_15       = c15["high"]
    low_15        = c15["low"]
    entry_trigger = round(high_15 + 1, 2)

    base.update({
        "15min_open"    : c15["open"],
        "15min_high"    : high_15,
        "15min_low"     : low_15,
        "15min_close"   : c15["close"],
        "entry_trigger" : entry_trigger,
    })

    # 3. Subsequent 5-min candles (strictly after 09:15)
    df5 = fetch_candles(kite, token, date_str, "5minute")
    time.sleep(API_DELAY)
    if df5.empty:
        return {**base, "error": "No 5-min data"}

    after_open = df5[df5["date"].dt.strftime("%H:%M") > MARKET_OPEN].copy()
    if after_open.empty:
        return {**base, "error": "No 5-min candles after 09:15"}

    # 4. Fibonacci retracement checks
    for lvl in FIB_LEVELS:
        price = fib_price(high_15, low_15, lvl)
        label = f"Fib_{int(lvl * 1000) / 10}%"
        base[f"{label}_price"]      = price
        base[f"{label}_hit_candle"] = first_touch(after_open, price)

    # 5. Long entry: first 5-min candle where high >= entry_trigger
    entry_row = None
    for _, row in after_open.iterrows():
        if row["high"] >= entry_trigger:
            entry_row = row
            break

    if entry_row is None:
        base.update({
            "entry_candle"  : "Entry not possible",
            "entry_price"   : None,
            "stop_loss"     : None,
            "sl_hit_candle" : "N/A",
            "max_gain"      : None,
            "max_gain_pct"  : None,
        })
        return base

    entry_time = entry_row["date"].strftime("%H:%M")
    stop_loss  = round(entry_row["low"] - 1, 2)   # SL = entry candle low - 1

    base.update({
        "entry_candle" : entry_time,
        "entry_price"  : entry_trigger,
        "stop_loss"    : stop_loss,
    })

    # 6. Walk candles after entry: track max high before SL is hit
    post_entry    = after_open[after_open["date"] > entry_row["date"]]
    sl_time       = None
    peak_price    = entry_trigger   # best price reached after entry

    for _, row in post_entry.iterrows():
        # Check SL first (gap-down open can breach SL immediately)
        if row["low"] <= stop_loss:
            sl_time = row["date"].strftime("%H:%M")
            # Price may have moved up intra-candle before hitting SL;
            # capture the high of this candle too if above current peak
            if row["high"] > peak_price:
                peak_price = row["high"]
            break
        # SL not hit — update running peak
        if row["high"] > peak_price:
            peak_price = row["high"]

    # If SL never hit, use EOD last candle close as indicative exit
    if sl_time:
        base["sl_hit_candle"] = sl_time
    else:
        last = after_open.iloc[-1]
        base["sl_hit_candle"] = f"Not hit (EOD @ {last['date'].strftime('%H:%M')})"
        if last["close"] > peak_price:
            peak_price = last["close"]

    max_gain     = round(peak_price - entry_trigger, 2)
    max_gain_pct = round((peak_price - entry_trigger) / entry_trigger * 100, 2)

    base["max_gain"]     = max_gain
    base["max_gain_pct"] = max_gain_pct
    return base


# ── Column order ──────────────────────────────────────────────────────────────

COLUMN_ORDER = [
    "date", "symbol",
    "15min_open", "15min_high", "15min_low", "15min_close",
    "Fib_38.2%_price", "Fib_38.2%_hit_candle",
    "Fib_50.0%_price", "Fib_50.0%_hit_candle",
    "Fib_61.8%_price", "Fib_61.8%_hit_candle",
    "entry_trigger", "entry_candle", "entry_price",
    "stop_loss", "sl_hit_candle",
    "max_gain", "max_gain_pct",
    "error",
]

# ── Color palette ─────────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color: str = "000000", bold: bool = False) -> Font:
    return Font(color=hex_color, bold=bold, name="Arial", size=10)

FILLS = {
    "header"        : fill("1F3864"),   # dark navy
    "range_candle"  : fill("D9E1F2"),   # light blue  — 15-min OHLC
    "fib_price"     : fill("FFF2CC"),   # light yellow — fib prices
    "fib_hit"       : fill("E2EFDA"),   # light green  — fib hit candles
    "entry_block"   : fill("FCE4D6"),   # light orange — entry columns
    "sl_block"      : fill("FFE699"),   # amber        — SL columns
    "gain_pos"      : fill("C6EFCE"),   # green        — positive gain
    "gain_neg"      : fill("FFC7CE"),   # red          — negative gain
    "fib_touched"   : fill("70AD47"),   # dark green   — fib actually hit
    "fib_not"       : fill("FF0000"),   # red          — not touched
    "entry_ok"      : fill("70AD47"),   # green        — entry taken
    "entry_no"      : fill("FF0000"),   # red          — entry not possible
    "sl_hit"        : fill("FF0000"),   # red          — SL hit
    "sl_not"        : fill("70AD47"),   # green        — SL not hit
    "error_row"     : fill("FFD7D7"),   # pale red     — error rows
}

FONTS = {
    "header"    : font("FFFFFF", bold=True),
    "fib_touch" : font("FFFFFF", bold=True),
    "entry_ok"  : font("FFFFFF", bold=True),
    "entry_no"  : font("FFFFFF", bold=True),
    "sl_hit"    : font("FFFFFF", bold=True),
    "sl_not"    : font("FFFFFF", bold=True),
    "gain_pos"  : font("276221"),
    "gain_neg"  : font("9C0006"),
}

thin_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ── Save to Excel with full color coding ──────────────────────────────────────

def save_excel(results: list, path: str):
    df   = pd.DataFrame(results)
    cols = [c for c in COLUMN_ORDER if c in df.columns]
    df   = df[cols]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]

        # ── Map column names to Excel column indices ──────────────────────────
        col_idx = {name: i + 1 for i, name in enumerate(cols)}

        # ── Header row styling ────────────────────────────────────────────────
        for col_num, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill      = FILLS["header"]
            cell.font      = FONTS["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border

        # ── Column group background fills (header + data) ─────────────────────
        group_fills = {
            **{c: "range_candle" for c in ["15min_open","15min_high","15min_low","15min_close"]},
            **{c: "fib_price"    for c in ["Fib_38.2%_price","Fib_50.0%_price","Fib_61.8%_price"]},
            **{c: "fib_hit"      for c in ["Fib_38.2%_hit_candle","Fib_50.0%_hit_candle","Fib_61.8%_hit_candle"]},
            **{c: "entry_block"  for c in ["entry_trigger","entry_candle","entry_price"]},
            **{c: "sl_block"     for c in ["stop_loss","sl_hit_candle"]},
        }

        n_rows = len(df)
        for col_name, fill_key in group_fills.items():
            if col_name not in col_idx:
                continue
            ci = col_idx[col_name]
            for r in range(2, n_rows + 2):
                cell = ws.cell(row=r, column=ci)
                cell.fill   = FILLS[fill_key]
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Apply borders to all remaining data cells ─────────────────────────
        for r in range(2, n_rows + 2):
            for ci in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=ci)
                if not cell.border or cell.border == Border():
                    cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Row-level logic-based color coding ────────────────────────────────
        fib_hit_cols = {
            "Fib_38.2%_hit_candle": col_idx.get("Fib_38.2%_hit_candle"),
            "Fib_50.0%_hit_candle": col_idx.get("Fib_50.0%_hit_candle"),
            "Fib_61.8%_hit_candle": col_idx.get("Fib_61.8%_hit_candle"),
        }
        entry_col  = col_idx.get("entry_candle")
        sl_col     = col_idx.get("sl_hit_candle")
        gain_col   = col_idx.get("max_gain")
        gainp_col  = col_idx.get("max_gain_pct")
        error_col  = col_idx.get("error")

        for r in range(2, n_rows + 2):
            row_data = {col: ws.cell(row=r, column=col_idx[col]).value
                        for col in cols if col in col_idx}
            is_error = bool(row_data.get("error"))

            # Error row: paint entire row
            if is_error:
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=r, column=ci).fill = FILLS["error_row"]
                continue

            # Fib hit cells
            for col_name, ci in fib_hit_cols.items():
                if ci is None:
                    continue
                cell = ws.cell(row=r, column=ci)
                val  = cell.value
                if val and val != "Not touched":
                    cell.fill = FILLS["fib_touched"]
                    cell.font = FONTS["fib_touch"]
                elif val == "Not touched":
                    cell.fill = FILLS["fib_not"]
                    cell.font = font("FFFFFF")

            # Entry candle cell
            if entry_col:
                cell = ws.cell(row=r, column=entry_col)
                if cell.value == "Entry not possible":
                    cell.fill = FILLS["entry_no"]
                    cell.font = FONTS["entry_no"]
                elif cell.value:
                    cell.fill = FILLS["entry_ok"]
                    cell.font = FONTS["entry_ok"]

            # SL hit cell
            if sl_col:
                cell = ws.cell(row=r, column=sl_col)
                val  = str(cell.value or "")
                if val == "N/A":
                    pass
                elif val.startswith("Not hit"):
                    cell.fill = FILLS["sl_not"]
                    cell.font = FONTS["sl_not"]
                elif val:
                    cell.fill = FILLS["sl_hit"]
                    cell.font = FONTS["sl_hit"]

            # Max gain cells
            for ci in [gain_col, gainp_col]:
                if ci is None:
                    continue
                cell = ws.cell(row=r, column=ci)
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0:
                        cell.fill = FILLS["gain_pos"]
                        cell.font = FONTS["gain_pos"]
                    else:
                        cell.fill = FILLS["gain_neg"]
                        cell.font = FONTS["gain_neg"]

        # ── Auto-size columns ─────────────────────────────────────────────────
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 3
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

        # ── Freeze top row ────────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Add a color legend sheet ──────────────────────────────────────────
        legend_data = [
            ("15-min OHLC columns",              "D9E1F2", "000000"),
            ("Fibonacci price levels",            "FFF2CC", "000000"),
            ("Fibonacci level — TOUCHED",         "70AD47", "FFFFFF"),
            ("Fibonacci level — Not touched",     "FF0000", "FFFFFF"),
            ("Entry columns background",          "FCE4D6", "000000"),
            ("Entry candle — TAKEN",              "70AD47", "FFFFFF"),
            ("Entry candle — NOT POSSIBLE",       "FF0000", "FFFFFF"),
            ("SL columns background",             "FFE699", "000000"),
            ("SL — HIT",                          "FF0000", "FFFFFF"),
            ("SL — NOT HIT (trade open at EOD)",  "70AD47", "FFFFFF"),
            ("Max gain — Positive",               "C6EFCE", "276221"),
            ("Max gain — Negative",               "FFC7CE", "9C0006"),
            ("Error row",                         "FFD7D7", "000000"),
        ]
        lws = writer.book.create_sheet("Legend")
        lws.column_dimensions["A"].width = 38
        lws.column_dimensions["B"].width = 16

        lws["A1"] = "Description"
        lws["B1"] = "Color sample"
        for cell in lws[1]:
            cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center")

        for i, (desc, bg, fg) in enumerate(legend_data, start=2):
            lws[f"A{i}"] = desc
            lws[f"A{i}"].font      = Font(name="Arial", size=10)
            lws[f"A{i}"].alignment = Alignment(vertical="center")
            lws[f"B{i}"].fill      = PatternFill("solid", fgColor=bg)
            lws[f"B{i}"].font      = Font(name="Arial", size=10, bold=True, color=fg)
            lws[f"B{i}"].value     = "Sample"
            lws[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")
            lws.row_dimensions[i].height = 18

    print(f"  Saved → {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Fibonacci retracement backtest via Kite API")
    parser.add_argument("--input",  default="Backtest-EarlyBloom.xlsx")
    parser.add_argument("--output", default="fib_results.xlsx")
    args = parser.parse_args()

    try:
        df_in = pd.read_excel(args.input)
    except FileNotFoundError:
        print(f"ERROR: '{args.input}' not found.")
        sys.exit(1)

    if not {"date", "symbol"}.issubset(df_in.columns):
        print("ERROR: Input file must have 'date' and 'symbol' columns.")
        sys.exit(1)

    df_in["date"] = pd.to_datetime(df_in["date"]).dt.strftime("%Y-%m-%d")

    print("Connecting to Kite API ...")
    kite = get_kite()
    print("Pre-loading NSE instrument list ...")
    kite.instruments(EXCHANGE)
    print("Ready.\n")

    results = []
    total   = len(df_in)

    for i, row in df_in.iterrows():
        symbol   = str(row["symbol"]).strip().upper()
        date_str = row["date"]
        print(f"[{i+1:>3}/{total}] {symbol:<20} {date_str} ...", end=" ", flush=True)

        result = analyse(kite, symbol, date_str)
        results.append(result)

        if "error" in result:
            print(f"ERROR → {result['error']}")
        else:
            print(
                f"Entry: {str(result.get('entry_candle')):<26}"
                f"SL: {str(result.get('sl_hit_candle')):<34}"
                f"MaxGain: {result.get('max_gain')} ({result.get('max_gain_pct')}%)"
            )

        # Save progress every 25 rows
        if (i + 1) % 25 == 0:
            save_excel(results, args.output)

    save_excel(results, args.output)

    # Print summary
    df_out  = pd.DataFrame(results)
    entered = df_out[df_out.get("entry_candle", pd.Series(dtype=str)) != "Entry not possible"]
    gains   = df_out["max_gain"].dropna() if "max_gain" in df_out.columns else pd.Series(dtype=float)

    print(f"\n{'─'*58}")
    print(f"Total rows    : {total}")
    print(f"Entry taken   : {len(entered)}")
    if len(gains):
        print(f"Win rate      : {int((gains > 0).sum())}/{len(gains)}")
        print(f"Avg max gain  : {gains.mean():.2f}")
        print(f"Best trade    : {gains.max():.2f}")
        print(f"Worst trade   : {gains.min():.2f}")
    print(f"{'─'*58}")


if __name__ == "__main__":
    main()